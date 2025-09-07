"""
Export functionality for expense data to various formats.
"""

import csv
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
from decimal import Decimal
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.chart import LineChart, PieChart, Reference
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportManager:
    """
    Manages data export to various formats (CSV, Excel, PDF, JSON).
    """
    
    def __init__(self):
        """Initialize export manager."""
        self.supported_formats = ['csv', 'json']
        
        if XLSX_AVAILABLE:
            self.supported_formats.append('xlsx')
        
        if PDF_AVAILABLE:
            self.supported_formats.append('pdf')
    
    def export_data(
        self, 
        data: Dict, 
        output_file: Path, 
        format_type: str = None,
        options: Optional[Dict] = None
    ) -> Tuple[bool, str]:
        """
        Export data to specified format.
        
        Args:
            data: Data to export
            output_file: Output file path
            format_type: Export format (auto-detected from extension if None)
            options: Export options
            
        Returns:
            Tuple of (success, message)
        """
        try:
            if format_type is None:
                format_type = output_file.suffix.lower().lstrip('.')
            
            if format_type not in self.supported_formats:
                return False, f"Unsupported format: {format_type}. Supported: {self.supported_formats}"
            
            options = options or {}
            
            # Ensure output directory exists
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            if format_type == 'csv':
                return self._export_csv(data, output_file, options)
            elif format_type == 'xlsx':
                return self._export_xlsx(data, output_file, options)
            elif format_type == 'pdf':
                return self._export_pdf(data, output_file, options)
            elif format_type == 'json':
                return self._export_json(data, output_file, options)
            else:
                return False, f"Format {format_type} not implemented"
                
        except Exception as e:
            error_msg = f"Export failed: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def _export_csv(self, data: Dict, output_file: Path, options: Dict) -> Tuple[bool, str]:
        """Export data to CSV format."""
        try:
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers
                headers = [
                    "Ano", "Mês", "Energia", "Água", "Internet", "Eventuais",
                    "Total", "Casa 1 Deve", "Casa 2 Deve", 
                    "Casa 1 Pagou", "Casa 2 Pagou", "Saldo Mês", "Saldo Acum."
                ]
                writer.writerow(headers)
                
                # Process data
                accumulated_balance = Decimal('0.00')
                
                for year_str, year_data in sorted(data.items()):
                    if not str(year_str).isdigit():
                        continue
                        
                    year = int(year_str)
                    months_order = [
                        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
                    ]
                    
                    for month in months_order:
                        if month not in year_data:
                            continue
                        
                        month_data = year_data[month]
                        results = month_data.get('results', {})
                        payments = month_data.get('payments', {})
                        
                        # Extract values
                        electricity = self._get_decimal_value(results.get('electricity', {}).get('total', 0))
                        water = self._get_decimal_value(results.get('recurring', {}).get('water_total', 0))
                        internet = self._get_decimal_value(results.get('recurring', {}).get('internet_total', 0))
                        
                        occasional_total = Decimal('0.00')
                        for expense in results.get('occasional', []):
                            occasional_total += self._get_decimal_value(expense.get('total', 0))
                        
                        total_expenses = self._get_decimal_value(results.get('total_expenses', 0))
                        casa1_should = self._get_decimal_value(results.get('casa1_should_pay', 0))
                        casa2_should = self._get_decimal_value(results.get('casa2_should_pay', 0))
                        casa1_paid = self._get_decimal_value(payments.get('casa1_paid', 0))
                        casa2_paid = self._get_decimal_value(payments.get('casa2_paid', 0))
                        month_balance = self._get_decimal_value(results.get('month_balance', 0))
                        
                        accumulated_balance += month_balance
                        
                        # Write row
                        row = [
                            year, month, float(electricity), float(water), float(internet),
                            float(occasional_total), float(total_expenses), float(casa1_should),
                            float(casa2_should), float(casa1_paid), float(casa2_paid),
                            float(month_balance), float(accumulated_balance)
                        ]
                        writer.writerow(row)
            
            return True, f"CSV exported successfully to {output_file}"
            
        except Exception as e:
            return False, f"CSV export failed: {str(e)}"
    
    def _export_xlsx(self, data: Dict, output_file: Path, options: Dict) -> Tuple[bool, str]:
        """Export data to Excel format with formatting and charts."""
        if not XLSX_AVAILABLE:
            return False, "openpyxl not available. Install with: pip install openpyxl"
        
        try:
            wb = openpyxl.Workbook()
            
            # Create main data sheet
            ws_data = wb.active
            ws_data.title = "Dados Anuais"
            
            # Create summary sheet
            ws_summary = wb.create_sheet("Resumo")
            
            # Create charts sheet
            ws_charts = wb.create_sheet("Gráficos")
            
            # Setup styles
            header_style = NamedStyle(name="header")
            header_style.font = Font(bold=True, color="FFFFFF")
            header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_style.alignment = Alignment(horizontal="center", vertical="center")
            
            currency_style = NamedStyle(name="currency")
            currency_style.number_format = 'R$ #,##0.00'
            
            # Add styles to workbook
            wb.add_named_style(header_style)
            wb.add_named_style(currency_style)
            
            # Populate data sheet
            self._populate_excel_data_sheet(ws_data, data, header_style, currency_style)
            
            # Populate summary sheet
            self._populate_excel_summary_sheet(ws_summary, data, header_style, currency_style)
            
            # Add charts
            self._add_excel_charts(ws_charts, data)
            
            # Save workbook
            wb.save(output_file)
            
            return True, f"Excel file exported successfully to {output_file}"
            
        except Exception as e:
            return False, f"Excel export failed: {str(e)}"
    
    def _populate_excel_data_sheet(self, ws, data, header_style, currency_style):
        """Populate Excel data sheet with expense data."""
        # Headers
        headers = [
            "Ano", "Mês", "Energia", "Água", "Internet", "Eventuais",
            "Total", "Casa 1 Deve", "Casa 2 Deve", 
            "Casa 1 Pagou", "Casa 2 Pagou", "Saldo Mês", "Saldo Acum."
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.style = header_style
        
        # Data rows
        row_num = 2
        accumulated_balance = Decimal('0.00')
        
        for year_str, year_data in sorted(data.items()):
            if not str(year_str).isdigit():
                continue
                
            year = int(year_str)
            months_order = [
                "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
            ]
            
            for month in months_order:
                if month not in year_data:
                    continue
                
                month_data = year_data[month]
                results = month_data.get('results', {})
                payments = month_data.get('payments', {})
                
                # Extract and convert values
                values = self._extract_month_values(results, payments)
                accumulated_balance += values['month_balance']
                
                # Populate row
                ws.cell(row=row_num, column=1).value = year
                ws.cell(row=row_num, column=2).value = month
                ws.cell(row=row_num, column=3).value = float(values['electricity'])
                ws.cell(row=row_num, column=4).value = float(values['water'])
                ws.cell(row=row_num, column=5).value = float(values['internet'])
                ws.cell(row=row_num, column=6).value = float(values['occasional'])
                ws.cell(row=row_num, column=7).value = float(values['total_expenses'])
                ws.cell(row=row_num, column=8).value = float(values['casa1_should'])
                ws.cell(row=row_num, column=9).value = float(values['casa2_should'])
                ws.cell(row=row_num, column=10).value = float(values['casa1_paid'])
                ws.cell(row=row_num, column=11).value = float(values['casa2_paid'])
                ws.cell(row=row_num, column=12).value = float(values['month_balance'])
                ws.cell(row=row_num, column=13).value = float(accumulated_balance)
                
                # Apply currency formatting to financial columns
                for col in range(3, 14):
                    ws.cell(row=row_num, column=col).style = currency_style
                
                # Color code balance
                balance_cell = ws.cell(row=row_num, column=12)
                if values['month_balance'] > 0:
                    balance_cell.font = Font(color="FF0000")  # Red
                elif values['month_balance'] < 0:
                    balance_cell.font = Font(color="008000")  # Green
                
                row_num += 1
        
        # Add totals row
        if row_num > 2:
            ws.cell(row=row_num, column=1).value = "TOTAL"
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            
            for col in range(3, 14):
                if col != 13:  # Skip accumulated balance column
                    col_letter = openpyxl.utils.get_column_letter(col)
                    formula = f"=SUM({col_letter}2:{col_letter}{row_num-1})"
                    cell = ws.cell(row=row_num, column=col)
                    cell.value = formula
                    cell.font = Font(bold=True)
                    cell.style = currency_style
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _populate_excel_summary_sheet(self, ws, data, header_style, currency_style):
        """Populate Excel summary sheet with statistics."""
        ws.cell(row=1, column=1).value = "RESUMO ANUAL"
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        
        row = 3
        
        # Calculate summary statistics
        stats = self._calculate_summary_statistics(data)
        
        summary_data = [
            ("Total de Anos:", len(data)),
            ("Total de Meses:", stats['total_months']),
            ("Gastos Totais:", f"R$ {stats['total_expenses']:.2f}"),
            ("Média Mensal:", f"R$ {stats['average_monthly']:.2f}"),
            ("Maior Gasto Mensal:", f"R$ {stats['highest_month']:.2f}"),
            ("Menor Gasto Mensal:", f"R$ {stats['lowest_month']:.2f}"),
            ("Saldo Final:", f"R$ {stats['final_balance']:.2f}")
        ]
        
        for label, value in summary_data:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).value = value
            row += 1
        
        # Monthly breakdown
        row += 2
        ws.cell(row=row, column=1).value = "BREAKDOWN POR CATEGORIA"
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        row += 2
        
        category_headers = ["Categoria", "Total", "Média Mensal", "% do Total"]
        for col, header in enumerate(category_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.style = header_style
        
        row += 1
        
        categories = [
            ("Energia", stats['total_electricity']),
            ("Água", stats['total_water']),
            ("Internet", stats['total_internet']),
            ("Eventuais", stats['total_occasional'])
        ]
        
        for category, total in categories:
            ws.cell(row=row, column=1).value = category
            ws.cell(row=row, column=2).value = float(total)
            ws.cell(row=row, column=2).style = currency_style
            
            if stats['total_months'] > 0:
                avg = total / stats['total_months']
                ws.cell(row=row, column=3).value = float(avg)
                ws.cell(row=row, column=3).style = currency_style
            
            if stats['total_expenses'] > 0:
                percentage = (total / stats['total_expenses']) * 100
                ws.cell(row=row, column=4).value = f"{percentage:.1f}%"
            
            row += 1
    
    def _add_excel_charts(self, ws, data):
        """Add charts to Excel workbook."""
        ws.cell(row=1, column=1).value = "GRÁFICOS E ANÁLISES"
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        
        # Note: Chart creation would require more complex data preparation
        # This is a placeholder for chart functionality
        ws.cell(row=3, column=1).value = "Gráficos serão implementados em versão futura"
        ws.cell(row=3, column=1).font = Font(italic=True)
    
    def _export_pdf(self, data: Dict, output_file: Path, options: Dict) -> Tuple[bool, str]:
        """Export data to PDF format."""
        if not PDF_AVAILABLE:
            return False, "reportlab not available. Install with: pip install reportlab"
        
        try:
            doc = SimpleDocTemplate(
                str(output_file),
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Container for PDF elements
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=colors.darkblue,
                alignment=1  # Center
            )
            
            # Title
            title = Paragraph("Relatório de Despesas Domésticas", title_style)
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Generate date
            generated_date = Paragraph(
                f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                styles['Normal']
            )
            story.append(generated_date)
            story.append(Spacer(1, 24))
            
            # Summary statistics
            summary_title = Paragraph("Resumo Executivo", styles['Heading2'])
            story.append(summary_title)
            
            stats = self._calculate_summary_statistics(data)
            summary_text = f"""
            <b>Total de meses com dados:</b> {stats['total_months']}<br/>
            <b>Gastos totais:</b> R$ {stats['total_expenses']:,.2f}<br/>
            <b>Média mensal:</b> R$ {stats['average_monthly']:,.2f}<br/>
            <b>Saldo final:</b> R$ {stats['final_balance']:,.2f}
            """
            
            summary_para = Paragraph(summary_text, styles['Normal'])
            story.append(summary_para)
            story.append(Spacer(1, 24))
            
            # Detailed data table
            table_title = Paragraph("Detalhamento Mensal", styles['Heading2'])
            story.append(table_title)
            story.append(Spacer(1, 12))
            
            # Prepare table data
            table_data = [
                ['Mês/Ano', 'Total', 'Casa 1', 'Casa 2', 'Saldo']
            ]
            
            for year_str, year_data in sorted(data.items()):
                if not str(year_str).isdigit():
                    continue
                
                year = int(year_str)
                months_order = [
                    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
                ]
                
                for month in months_order:
                    if month not in year_data:
                        continue
                    
                    month_data = year_data[month]
                    results = month_data.get('results', {})
                    
                    values = self._extract_month_values(results, month_data.get('payments', {}))
                    
                    table_data.append([
                        f"{month}/{year}",
                        f"R$ {values['total_expenses']:,.2f}",
                        f"R$ {values['casa1_should']:,.2f}",
                        f"R$ {values['casa2_should']:,.2f}",
                        f"R$ {values['month_balance']:,.2f}"
                    ])
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            
            # Build PDF
            doc.build(story)
            
            return True, f"PDF exported successfully to {output_file}"
            
        except Exception as e:
            return False, f"PDF export failed: {str(e)}"
    
    def _export_json(self, data: Dict, output_file: Path, options: Dict) -> Tuple[bool, str]:
        """Export data to JSON format."""
        try:
            # Convert Decimal objects to strings for JSON serialization
            serializable_data = self._prepare_data_for_json(data)
            
            # Add export metadata
            export_data = {
                'metadata': {
                    'exported_at': datetime.now().isoformat(),
                    'format_version': '2.0',
                    'total_years': len(data),
                    'export_options': options
                },
                'data': serializable_data
            }
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
            
            return True, f"JSON exported successfully to {output_file}"
            
        except Exception as e:
            return False, f"JSON export failed: {str(e)}"
    
    def _prepare_data_for_json(self, data: Dict) -> Dict:
        """Prepare data for JSON serialization."""
        def convert_value(obj):
            if isinstance(obj, Decimal):
                return str(obj)
            elif isinstance(obj, dict):
                return {k: convert_value(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_value(item) for item in obj]
            else:
                return obj
        
        return convert_value(data)
    
    def _get_decimal_value(self, value: Any) -> Decimal:
        """Convert value to Decimal safely."""
        if isinstance(value, Decimal):
            return value
        elif isinstance(value, (int, float, str)):
            try:
                return Decimal(str(value))
            except:
                return Decimal('0.00')
        else:
            return Decimal('0.00')
    
    def _extract_month_values(self, results: Dict, payments: Dict) -> Dict:
        """Extract and convert month values to Decimal."""
        electricity = self._get_decimal_value(results.get('electricity', {}).get('total', 0))
        water = self._get_decimal_value(results.get('recurring', {}).get('water_total', 0))
        internet = self._get_decimal_value(results.get('recurring', {}).get('internet_total', 0))
        
        occasional_total = Decimal('0.00')
        for expense in results.get('occasional', []):
            occasional_total += self._get_decimal_value(expense.get('total', 0))
        
        return {
            'electricity': electricity,
            'water': water,
            'internet': internet,
            'occasional': occasional_total,
            'total_expenses': self._get_decimal_value(results.get('total_expenses', 0)),
            'casa1_should': self._get_decimal_value(results.get('casa1_should_pay', 0)),
            'casa2_should': self._get_decimal_value(results.get('casa2_should_pay', 0)),
            'casa1_paid': self._get_decimal_value(payments.get('casa1_paid', 0)),
            'casa2_paid': self._get_decimal_value(payments.get('casa2_paid', 0)),
            'month_balance': self._get_decimal_value(results.get('month_balance', 0))
        }
    
    def _calculate_summary_statistics(self, data: Dict) -> Dict:
        """Calculate summary statistics for the data."""
        stats = {
            'total_months': 0,
            'total_expenses': Decimal('0.00'),
            'total_electricity': Decimal('0.00'),
            'total_water': Decimal('0.00'),
            'total_internet': Decimal('0.00'),
            'total_occasional': Decimal('0.00'),
            'final_balance': Decimal('0.00'),
            'highest_month': Decimal('0.00'),
            'lowest_month': Decimal('999999.99'),
            'average_monthly': Decimal('0.00')
        }
        
        monthly_totals = []
        
        for year_data in data.values():
            for month_data in year_data.values():
                results = month_data.get('results', {})
                values = self._extract_month_values(results, month_data.get('payments', {}))
                
                stats['total_months'] += 1
                stats['total_expenses'] += values['total_expenses']
                stats['total_electricity'] += values['electricity']
                stats['total_water'] += values['water']
                stats['total_internet'] += values['internet']
                stats['total_occasional'] += values['occasional']
                stats['final_balance'] += values['month_balance']
                
                monthly_totals.append(values['total_expenses'])
                
                if values['total_expenses'] > stats['highest_month']:
                    stats['highest_month'] = values['total_expenses']
                
                if values['total_expenses'] < stats['lowest_month'] and values['total_expenses'] > 0:
                    stats['lowest_month'] = values['total_expenses']
        
        if stats['total_months'] > 0:
            stats['average_monthly'] = stats['total_expenses'] / stats['total_months']
        
        if stats['lowest_month'] == Decimal('999999.99'):
            stats['lowest_month'] = Decimal('0.00')
        
        return stats
    
    def get_export_options(self, format_type: str) -> Dict:
        """Get available export options for a specific format."""
        base_options = {
            'include_metadata': True,
            'date_format': 'iso',
            'currency_format': 'decimal'
        }
        
        if format_type == 'csv':
            return {
                **base_options,
                'delimiter': ',',
                'encoding': 'utf-8',
                'include_totals': True
            }
        elif format_type == 'xlsx':
            return {
                **base_options,
                'include_charts': True,
                'include_summary': True,
                'apply_formatting': True,
                'auto_width': True
            }
        elif format_type == 'pdf':
            return {
                **base_options,
                'page_size': 'A4',
                'include_summary': True,
                'include_charts': False
            }
        elif format_type == 'json':
            return {
                **base_options,
                'indent': 2,
                'sort_keys': True,
                'include_export_metadata': True
            }
        else:
            return base_options